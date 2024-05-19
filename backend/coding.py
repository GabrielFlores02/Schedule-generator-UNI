from openpyxl import Workbook
from openpyxl.styles import PatternFill,Color,Font,Border,Side

from datetime import datetime

def esta_dentro_del_intervalo(horario, intervalo):
    # Dividir la cadena en inicio y fin
    inicio_str, fin_str = horario.split('-')
    
    # Convertir las cadenas a objetos de tiempo
    inicio = datetime.strptime(inicio_str, "%H:%M").time()
    fin = datetime.strptime(fin_str, "%H:%M").time()
    
    # Convertir las cadenas de intervalo a objetos de tiempo
    inicio_intervalo, fin_intervalo = map(lambda x: datetime.strptime(x, "%H:%M").time(), intervalo.split('-'))
    
    # Verificar si el intervalo está dentro del rango designado
    return inicio_intervalo <= inicio <= fin_intervalo and inicio_intervalo <= fin <= fin_intervalo

def crear_horarios(data,colores,horas_limite):
    #Columnas de nuestro DATASET:
    codigo=data['A']
    nombrecursou=data['B']
    seccion=data['C']
    tipo=data['D']
    dia=data['E']
    hora=data['F']
    docente=data['G']
    departamento=data['H']
    modalidad=data['I']

    #Crear la primera pagina
    book = Workbook()
    sheet = book.active
    #Poniendo titulo
    sheet.title="horario1"
    sheet['D1']='    Horario'
    #Creando las horas
    for hour in range(3,18):
        sheet[f'A{hour}']=f'{hour+5} a {hour+6}'
    #Creando los dias
    sheet['B2']='Lunes'
    sheet['C2']='Martes'
    sheet['D2']='Miercoles'
    sheet['E2']='Jueves'
    sheet['F2']='Viernes'
    sheet['G2']='Sabado'
    sheet['A20']='Curso:'
    sheet['A21']='Código:'
    sheet['A22']='Sección:'

    # creamos el borde doble   otro dia :v
    #borde_doble = Side(border_style="double")
    #sheet["A1"].border = Border(top=borde_doble,
    #                        right=borde_doble,
    #                        bottom=borde_doble,
    #                        left=borde_doble)

    #Ancho de las columnas
    colect23=['A','B','C','D','E','F','G']
    for aaa in colect23:
        sheet.column_dimensions[aaa].width= 20

    #Colores: -------------> Agrega lo colores si quieres y reemplaza mas adelante el nombre nuevo
    ColorFill1 = PatternFill(patternType='solid',
                                        fgColor=colores[0])
    ColorFill2 = PatternFill(patternType='solid',
                                        fgColor=colores[1])
    ColorFill3 = PatternFill(patternType='solid',
                                        fgColor=colores[2])
    ColorFill4 = PatternFill(patternType='solid',
                                        fgColor=colores[3])
    ColorFill5 = PatternFill(patternType='solid',
                                        fgColor=colores[4])
    ColorFill6 = PatternFill(patternType='solid',
                                        fgColor=colores[5])
    ColorFill7 = PatternFill(patternType='solid',
                                        fgColor=colores[6])
    #Analisis General:
    origencod=codigo[1].value
    cursoactual = origencod
    intervalocurson2=1
    #analizamos el numero de cursos
    mi_lista_sin_duplicados = []
    for elemento in codigo:
        if elemento.value not in mi_lista_sin_duplicados:
            print(elemento.value)
            mi_lista_sin_duplicados.append(elemento.value)
    if None in mi_lista_sin_duplicados:
        mi_lista_sin_duplicados.remove(None)
    ncursos=len(mi_lista_sin_duplicados)-1 #(N° de cursos + EL ULTIMO) - El titulo Codigo
    jk=0 #variable para cambiar la letra del curso
    print("Numero de cursos : ", ncursos)
    for nc in range(ncursos):

        #Calcular la lineas a tratar por curso (intervalo: [intervalocurson1,intervalocurson2])
        intervalocurson1=intervalocurson2
        print(f'Desde {intervalocurson1}')
        while codigo[intervalocurson2].value == cursoactual:
            intervalocurson2+=1
        print(f'Hasta {intervalocurson2}')

        #Calcular el numero de libros actual antes de empezar a analizar el curso
        nombressheets=book.sheetnames
        nsheets=len(nombressheets)
        print('Numero de sheets a analizar:')
        print(nsheets)

        #Asignando la columna donde irá el nombre del curso
        letras=['B','C','D','E','F','G','H','I']
        letracurso=letras[jk]
        jk+=1
        isheet=0
        for n in nombressheets:   #Para cada sheet actual origen
            isheet=isheet+1
            sheet=book[n]
            hayespacio=False

            for i in range (intervalocurson1,intervalocurson2):  #Para cada linea en el intervalo del curso
                print(f"proceso{i}")
                #Si estamos en la primera linea entonces de cada seccion del curso
                if seccion[i].value != seccion[i-1].value:
                    #Lineas de la seccion
                    seccionlines=1
                    l=i
                    while seccion[l].value ==seccion[l+1].value:
                        seccionlines+=1
                        l+=1

                    print('Lineas por seccion:')
                    print(seccionlines)

                    #Verificamos si hay espacio para la seccion?
                    hayespacio=True
                    espacio=True
                    les=i
                    for s in range(seccionlines):
                        di=dia[les].value
                        hor=hora[les].value
                        if(di=='LU'):
                            col='B'
                        if(di=='MA'):
                            col='C'
                        if(di=='MI'):
                            col='D'
                        if(di=='JU'):
                            col='E'
                        if(di=='VI'):
                            col='F'
                        if(di=='SA'):
                            col='G'     
                        if(hor=='08:00-10:00'):
                            fil=[3,4]
                        if(hor=='08:00-11:00'):
                            fil=[3,4,5]
                        if(hor=='09:00-11:00'):
                            fil=[4,5]
                        if(hor=='09:00-12:00'):
                            fil=[4,5,6]
                        for j in range(10,22):
                            cont=2
                            for k in range(j+2,j+4):
                                if(hor==f'{j}:00-{k}:00'):
                                    if(cont==2):
                                        fil=[j-5,j-4]
                                    if(cont==3):
                                        fil=[j-5,j-4,j-3]
                                cont+=1
                        tam=len(fil)
                        if(tam==2):
                            if(sheet[f'{col}{fil[0]}'].value!=None or sheet[f'{col}{fil[1]}'].value!=None):
                                espacio=False
                        if(tam==3):
                            if(sheet[f'{col}{fil[0]}'].value!=None or sheet[f'{col}{fil[1]}'].value!=None or sheet[f'{col}{fil[2]}'].value!=None ):
                                espacio=False
                        hayespacio = hayespacio and espacio
                        les+=1

                        if esta_dentro_del_intervalo(hor, horas_limite):
                            print(f'El horario {hor} está dentro del intervalo designado.')
                            pass
                        else:
                            print(f'El horario {hor} NO está dentro del intervalo designado.')
                            hayespacio=False

                #Si hay espacio en el curso en el sheet --> procede xd
                if(hayespacio==True):

                    #Creamos el sheet copia donde se trabaja, dejando al sheet origen como modelo a analizar
                    
                    if seccion[i].value !=seccion[i-1].value:
                        sheetcopia = book.copy_worksheet(sheet)#---------------------------------------------------------------------------------
                        sheetcopia.title=f'horario{nsheets+1}'
                        nsheets+=1

                    #Sacamos los valores de la lineas respecctiva
                    cod=codigo[i].value
                    ncu=nombrecursou[i].value
                    sec=seccion[i].value
                    tip=tipo[i].value
                    di=dia[i].value
                    hor=hora[i].value
                    doc=docente[i].value
                    dep=departamento[i].value
                    mod=modalidad[i].value
                    
                    #ubicaion por columna
                    if(di=='LU'):
                        col='B'
                    if(di=='MA'):
                        col='C'
                    if(di=='MI'):
                        col='D'
                    if(di=='JU'):
                        col='E'
                    if(di=='VI'):
                        col='F'
                    if(di=='SA'):
                        col='G'
                    #ubicacion por filas
                    if(hor=='08:00-10:00'):
                        fil=[3,4]
                    if(hor=='08:00-11:00'):
                        fil=[3,4,5]
                    if(hor=='09:00-11:00'):
                        fil=[4,5]
                    if(hor=='09:00-12:00'):
                        fil=[4,5,6]
                    for j in range(10,22):
                        cont=2
                        for k in range(j+2,j+4):
                            if(hor==f'{j}:00-{k}:00'):
                                if(cont==2):
                                    fil=[j-5,j-4]
                                if(cont==3):
                                    fil=[j-5,j-4,j-3]
                            cont+=1
                    print(f'hora: {hor} ---> linea: {i}')
                    print(f'fila: {fil} and columna: {col}\n')
                    #Ponemos en el horario
                    tam=len(fil)
                    if(tam==2):
                            sheetcopia[f'{col}{fil[0]}']=cod
                            sheetcopia[f'{col}{fil[1]}']=tip
                            if(nc==0):  #Para el 1 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill1
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill1
                            if(nc==1):  #Para el 2 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill2
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill2
                            if(nc==2):  #Para el 3 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill3
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill3
                            if(nc==3):  #Para el 4 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill4
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill4
                            if(nc==4):  #Para el 5 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill5
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill5
                            if(nc==5):  #Para el 6 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill6
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill6
                            if(nc==6):  #Para el 7 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill7
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill7


                    if(tam==3):
                            sheetcopia[f'{col}{fil[0]}']=cod
                            sheetcopia[f'{col}{fil[1]}']=tip
                            sheetcopia[f'{col}{fil[2]}']=' '
                            if(nc==0):  #Para el 1 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill1
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill1
                                sheetcopia[f'{col}{fil[2]}'].fill=ColorFill1
                            if(nc==1):  #Para el 2 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill2
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill2
                                sheetcopia[f'{col}{fil[2]}'].fill=ColorFill2
                            if(nc==2):  #Para el 3 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill3
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill3
                                sheetcopia[f'{col}{fil[2]}'].fill=ColorFill3
                            if(nc==3):  #Para el 4 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill4
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill4
                                sheetcopia[f'{col}{fil[2]}'].fill=ColorFill4
                            if(nc==4):  #Para el 5 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill5
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill5
                                sheetcopia[f'{col}{fil[2]}'].fill=ColorFill5
                            if(nc==5):  #Para el 6 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill6
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill6
                                sheetcopia[f'{col}{fil[2]}'].fill=ColorFill6
                            if(nc==6):  #Para el 7 curso coloreo
                                sheetcopia[f'{col}{fil[0]}'].fill=ColorFill7
                                sheetcopia[f'{col}{fil[1]}'].fill=ColorFill7
                                sheetcopia[f'{col}{fil[2]}'].fill=ColorFill7

                    #Ponemos los datos de la parte de abajo
                    sheetcopia[f'{letracurso}20']=ncu
                    sheetcopia[f'{letracurso}21']=cod
                    sheetcopia[f'{letracurso}22']=sec
                    if(nc==0):#Para el 1 curso coloreo codigo y seccion
                        sheetcopia[f'{letracurso}21'].fill=ColorFill1
                        sheetcopia[f'{letracurso}22'].fill=ColorFill1
                    if(nc==1):#Para el 2 curso coloreo codigo y seccion
                        sheetcopia[f'{letracurso}21'].fill=ColorFill2
                        sheetcopia[f'{letracurso}22'].fill=ColorFill2                    
                    if(nc==2):#Para el 3 curso coloreo codigo y seccion
                        sheetcopia[f'{letracurso}21'].fill=ColorFill3
                        sheetcopia[f'{letracurso}22'].fill=ColorFill3                   
                    if(nc==3):#Para el 4 curso coloreo codigo y seccion
                        sheetcopia[f'{letracurso}21'].fill=ColorFill4
                        sheetcopia[f'{letracurso}22'].fill=ColorFill4                    
                    if(nc==4):#Para el 5 curso coloreo codigo y seccion
                        sheetcopia[f'{letracurso}21'].fill=ColorFill5
                        sheetcopia[f'{letracurso}22'].fill=ColorFill5
                    if(nc==5):#Para el 6 curso coloreo codigo y seccion
                        sheetcopia[f'{letracurso}21'].fill=ColorFill6
                        sheetcopia[f'{letracurso}22'].fill=ColorFill6                    
                    if(nc==6):#Para el 7 curso coloreo codigo y seccion
                        sheetcopia[f'{letracurso}21'].fill=ColorFill7
                        sheetcopia[f'{letracurso}22'].fill=ColorFill7
                    

                    if(tip=='T'):
                        sheetcopia[f'{letracurso}23']='T'
                        sheetcopia[f'{letracurso}24']=doc
                        sheetcopia[f'{letracurso}25']=dep
                    if(tip=='P'):
                        sheetcopia[f'{letracurso}26']='P'
                        sheetcopia[f'{letracurso}27']=doc
                        sheetcopia[f'{letracurso}28']=dep
                    if(tip=='L'):
                        sheetcopia[f'{letracurso}29']='L'
                        sheetcopia[f'{letracurso}30']=doc
                        sheetcopia[f'{letracurso}31']=dep
                    if(tip=='P/L'):
                        sheetcopia[f'{letracurso}26']='P/L'
                        sheetcopia[f'{letracurso}27']=doc
                        sheetcopia[f'{letracurso}28']=dep

                    book.save('Horarios.xlsx')
                else:
                    print('No hay espacio...')

        #Borrado de sheets que no sirven para 5 cursos:
        nombressheets=book.sheetnames
        if(jk==1):     #Para el 1 curso borramos horarios q no sirven
            for ns in nombressheets:
                sheet=book[ns]
                if sheet[f'{letras[0]}21'].value==None:
                    book.remove (book[ns])
        if(jk==2):  #Para el 2 curso borramos horarios q no sirven
            for ns in nombressheets:
                sheet=book[ns]
                if ((sheet[f'{letras[0]}21'].value==None) or (sheet[f'{letras[1]}21'].value==None)):
                    book.remove (book[ns])
        if(jk==3):  #Para el 3 curso borramos horarios q no sirven
            for ns in nombressheets:
                sheet=book[ns]
                if ((sheet[f'{letras[0]}21'].value==None) or (sheet[f'{letras[1]}21'].value==None)or (sheet[f'{letras[2]}21'].value==None)):
                    book.remove (book[ns])
        if(jk==4):  #Para el 4 curso borramos horarios q no sirven
            for ns in nombressheets:
                sheet=book[ns]
                if ((sheet[f'{letras[0]}21'].value==None) or (sheet[f'{letras[1]}21'].value==None)or (sheet[f'{letras[2]}21'].value==None)or (sheet[f'{letras[3]}21'].value==None)):
                    book.remove (book[ns])
        if(jk==5):  #Para el 5 curso borramos horarios q no sirven
            for ns in nombressheets:
                sheet=book[ns]
                if ((sheet[f'{letras[0]}21'].value==None) or (sheet[f'{letras[1]}21'].value==None)or (sheet[f'{letras[2]}21'].value==None)or (sheet[f'{letras[3]}21'].value==None)or (sheet[f'{letras[4]}21'].value==None)):
                    book.remove (book[ns])
        if(jk==6):  #Para el 6 curso borramos horarios q no sirven
            for ns in nombressheets:
                sheet=book[ns]
                if ((sheet[f'{letras[0]}21'].value==None) or (sheet[f'{letras[1]}21'].value==None)or (sheet[f'{letras[2]}21'].value==None)or (sheet[f'{letras[3]}21'].value==None)or (sheet[f'{letras[4]}21'].value==None)or (sheet[f'{letras[5]}21'].value==None)):
                    book.remove (book[ns])
        if(jk==7):  #Para el 7 curso borramos horarios q no sirven
            for ns in nombressheets:
                sheet=book[ns]
                if ((sheet[f'{letras[0]}21'].value==None) or (sheet[f'{letras[1]}21'].value==None)or (sheet[f'{letras[2]}21'].value==None)or (sheet[f'{letras[3]}21'].value==None)or (sheet[f'{letras[4]}21'].value==None)or (sheet[f'{letras[5]}21'].value==None)or (sheet[f'{letras[6]}21'].value==None)):
                    book.remove (book[ns])
                    
        book.save('Horarios.xlsx')
        cursoactual=codigo[intervalocurson2].value
        print(codigo[intervalocurson2].value)

    #Cambiamos el nombre de los sheetsnames
    nombressheets=book.sheetnames
    i=1
    for n in nombressheets:   #Para cada sheet actual origen
        sheet=book[n]
        sheet.title=f'Horario {i}'
        i+=1
    book.save('Horarios.xlsx')

